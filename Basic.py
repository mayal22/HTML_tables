from codecs import ignore_errors
from mailbox import NoSuchMailboxError
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains as AC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException ,NoSuchElementException
from openpyxl import Workbook
class Initials:
    opt = webdriver.ChromeOptions()
    driver = webdriver.Chrome(executable_path="C:\\Data driven testing\\chromedriver.exe")
    @classmethod
    def change_driver(cls,browser):
        browser = browser.upper()
        if  browser == "FIREFOX":
            cls.driver = webdriver.Firefox(executable_path="C:\\Data driven testing\\")
        elif browser == "SAFARI":
            cls.driver = webdriver.Safari(executable_path="C:\\Data driven testing\\")
        elif browser == "EDGE":
            cls.driver = webdriver.Edge(executable_path="C:\\Data driven testing\\")
        else:
            cls.driver = webdriver.Opera(executable_path="C:\\Data driven testing\\")
        return cls.driver
    @classmethod
    def hit_url(cls,url):
        return cls.driver.get(url)
    @classmethod
    def get_title(cls):
        return cls.driver.title
    @classmethod
    def scroll_down(cls,horizontal_pixels,vertical_pixels):
        script = f"window.scrollBy({horizontal_pixels},{vertical_pixels})"
        cls.driver.execute_script(script)
        return script
    @classmethod
    def hover_click(cls,move_locator,click_locator,move_name,click_name):
        move_locator = move_locator.upper()
        click_locator= click_locator.upper()
        action = AC(cls.driver)
        moving_object = cls.driver.find_element(By.move_locator,move_name)
        clicking_object = cls.driver.find_element(By.click_locator,click_name)
        action.move_to_element(moving_object)
        action.click(clicking_object)
        action.perform()
    @staticmethod
    def check_locator(locator,locator_name):
        if locator == "id":
            return Initials.driver.find_element(By.ID,locator_name)
        elif locator == "css selector":
            return Initials.driver.find_element(By.CSS_SELECTOR,locator_name)
        elif locator == "xpath":
            return Initials.driver.find_element(By.XPATH,locator_name)
        elif locator == "name":
            return Initials.driver.find_element(By.NAME,locator_name)
        elif locator == "tag name":
            return Initials.driver.find_element(By.TAG_NAME,locator_name)
        else:
            return Initials.driver.find_element(By.CLASS_NAME,locator_name)
    @staticmethod
    def check_locators(locator,locator_name):
        if locator == "id":
            return Initials.driver.find_elements(By.ID,locator_name)
        elif locator == "css selector":
            return Initials.driver.find_elements(By.CSS_SELECTOR,locator_name)
        elif locator == "xpath":
            return Initials.driver.find_elements(By.XPATH,locator_name)
        elif locator == "name":
            return Initials.driver.find_elements(By.NAME,locator_name)
        elif locator == "tag name":
            return Initials.driver.find_elements(By.TAG_NAME,locator_name)
        else:
            return Initials.driver.find_elements(By.CLASS_NAME,locator_name)
    @staticmethod
    def click(locator,locator_name):
        final_locator = None
        if locator == "id":
            final_locator = By.ID
        elif locator == "css selector":
            final_locator = By.CSS_SELECTOR
        elif locator == "xpath":
            final_locator = By.XPATH
        elif locator == "name":
            final_locator = By.NAME
        elif locator == "tag name":
            final_locator == By.TAG_NAME
        else:
            final_locator == By.CLASS_NAME
        waits = WebDriverWait(Initials.driver,10).until(EC.element_to_be_clickable((final_locator,locator_name)))
        waits.click()
    @staticmethod
    def insert_data(locator,locator_name,data):
        Initials.check_locator(locator,locator_name).send_keys(data)
    @staticmethod
    def current_url():
        return Initials.driver.current_url
    @staticmethod
    def get_text_elementbyelements(locator,locator_name,locators,locators_name):
        one = Initials.check_locator(locator,locator_name)
        myvalues = one.find_elements(Initials.get_locator(locators),locators_name)
        for each in myvalues:
            yield each.text
    @staticmethod
    def get_locator(locator):
        if locator =="css selector":
            return By.CSS_SELECTOR
        elif locator == "tag name":
            return By.TAG_NAME
        elif locator == "xpath":
            return By.XPATH
        elif locator == "name":
            return By.NAME
        elif locator == "class name":
            return By.CLASS_NAME
        else:
            return By.ID
    @staticmethod
    def numbers_of_elements(locator,locator_name):
        counts = Initials.check_locators(locator,locator_name)
        return len(counts)
    @staticmethod
    def element_visible(locator,locator_name):
        wait = WebDriverWait(Initials.driver,10).until(EC.visibility_of(Initials.check_locator(locator,locator_name)))
        Initials.check_locator(locator,locator_name).click()
    @staticmethod
    def element_located(locator,locator_name):
        fluent = WebDriverWait(Initials.driver,10,4,None)
        fluent.until(EC.presence_of_element_located(Initials.check_locator(locator,locator_name)))
    @staticmethod
    def url_wait(url_to_check):
        try:
            wait = WebDriverWait(Initials.driver,10).until(EC.url_changes(url_to_check))
        except TimeoutException:
            return f"Url not changes from {Initials.current_url()} to {url_to_check}"
    @staticmethod
    def title_present_wait(title_included):
        try:
            wait = WebDriverWait(Initials.driver,6)
            wait.until(EC.title_contains(title_included))
        except TimeoutException:
            return f"not found {title_included} after wait"
    @staticmethod
    def error_alerts(locator,locator_name):
        try:
            wait = WebDriverWait(Initials.driver,20).until(EC.presence_of_element_located((Initials.get_locator(locator),locator_name)))
            finals = Initials.driver.find_element(Initials.get_locator(locator),locator_name).text
        except TimeoutException as e:
            finals = "no error found"
        return finals
    @classmethod
    def New_Window(cls,link,window_count):
        cls.driver.execute_script("window.open('');")
        cls.driver.switch_to.window(cls.driver.window_handles[window_count])
        cls.hit_url(link)
    @staticmethod
    def tableData(tablefunc):
        def capturing_table_data(locator,locator_name,sheet_name):
            tables = tablefunc(locator,locator_name,sheet_name)
            c = 1
            r = 2
            heading = []
            dataset = []
            wb = Workbook()
            file_name = "C:\\Data driven testing\\Utilities\\tables_data.xlsx"
            sheet = wb.create_sheet(sheet_name)
            try:
                for each in tables:
                    headers = each.find_elements(By.TAG_NAME,"thead")
                    for each in headers:
                        header_data = each.find_elements(By.TAG_NAME,"tr")
                    for each in header_data:
                        theads = each.find_elements(By.TAG_NAME,"th")
                        for each in theads:
                            heading.append(each.text)
            except NoSuchElementException:
                print("sorry! something has wrong with it please check elements")
            for i in range(1,len(heading)+1):
                        sheet.cell(1,i,heading[i-1])
            try:
                for each in tables:
                    body = each.find_elements(By.TAG_NAME,"tbody")
                    for each in body:
                        rows = each.find_elements(By.TAG_NAME,"tr")
                    for each in rows:
                        data = each.find_elements(By.TAG_NAME,"td")
                        for each in data:
                            sheet.cell(r,c,each.text)
                            c+=1
                        r+=1
                        c = 1
                wb.save(file_name)
                return "successfull"
            except:
                print("sorry we cant find element in the process, please check the locators")
                print(dataset , heading)
        return capturing_table_data

    @staticmethod
    @tableData
    def Capture_HTML_Table(locator,locator_name,sheet_name):
        '''It is to load the table data in a excel sheet
        named as 'tables_data' and the sheets name as user given in argument
        '''
        table = Initials.check_locators(locator,locator_name)
        return table
    @staticmethod
    def text_wait(locator,locator_name,text):
        waits = WebDriverWait(Initials.driver,15)
        waits.until(EC.text_to_be_present_in_element((Initials.get_locator(locator),locator_name),text))
    @staticmethod
    def Selection_List(locator,locator_name,option):
        '''Function used to click the HTML of Selection Menu based on option provided by 
        user , it takes option as index , but if option value is True then it takes
        option as value
        '''
        select_web = Initials.check_locator(locator,locator_name)
        selection_menu = Select(select_web)
        all_option = selection_menu.options
        try:
            # try:
                if type(option) == str:
                    selection_menu.select_by_value(option)
                # try:
                    if option in range(0,len(all_option)):
                        selection_menu.select_by_index(option)
                # except LookupError as L:

                # else:
                #     selection_menu.select_by_value(option)
        except NoSuchElementException as N:
            N.msg="You given invalid values"
            print(N.msg)
            print("Available Options to write:")
            for each in all_option:
                print(each.text)
            print(f'your Given value : {option}')
        


                